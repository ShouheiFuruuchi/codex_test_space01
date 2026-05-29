from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict, deque
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "OUTPUT"
SUMMARY_SHEET = "社販集計表"
ALADDIN_SHEET = "アラジンデータ"
RESULT_SHEET = "照合結果"
DETAIL_SHEET = "照合差異明細"
PAIR_SHEET = "照合データ一覧"
ALADDIN_EMPLOYEE_COL = 1  # A
ALADDIN_PRODUCT_CODE_COL = 8  # H
ALADDIN_PRODUCT_NAME_COL = 9  # I
ALADDIN_COLOR_COL = 12  # L
ALADDIN_SIZE_COL = 14  # N
ALADDIN_COST_COL = 22  # V
ALADDIN_WHOLESALE_ADD_COL = 24  # X
ALADDIN_JAN_COL = 87  # CI


def normalize_digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def normalize_jan(value: Any) -> str:
    digits = normalize_digits(value)
    if len(digits) > 13:
        return digits[-13:]
    if len(digits) == 14 and digits.startswith("0"):
        return digits[1:]
    return digits


def parse_number(value: Any) -> float:
    text = str(value or "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_aladdin_text(value: Any) -> str:
    return str(value or "").replace("_x001F_", "").strip()


def parse_yyyymmdd(value: Any) -> tuple[Any, Any]:
    digits = normalize_digits(value)
    if len(digits) >= 6:
        year = int(digits[:4])
        month = int(digits[4:6])
        if 2000 <= year <= 2100 and 1 <= month <= 12:
            return year, month
    return "", ""


def latest_workbook() -> Path:
    files = sorted(
        [path for path in OUTPUT_DIR.glob("社販集計表_*.xlsx") if "照合完了" not in path.stem],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not files:
        raise FileNotFoundError("OUTPUT内に 社販集計表_*.xlsx が見つかりません。")
    return files[0]


def compared_output_path(path: Path) -> Path:
    if "照合完了" in path.stem:
        return path
    return path.with_name(f"{path.stem}_照合完了{path.suffix}")


def read_summary(ws) -> tuple[dict[str, str], dict[str, Counter[str]], list[dict[str, Any]], list[str]]:
    headers = [cell.value for cell in ws[1]]
    header_index = {name: index + 1 for index, name in enumerate(headers)}
    employee_col = header_index["社員CD"]
    name_col = header_index["氏名"]
    jan_col = header_index["JANコード"]

    employee_names: dict[str, str] = {}
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    rows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        employee_cd = normalize_digits(ws.cell(row, employee_col).value)[-4:]
        jan = normalize_jan(ws.cell(row, jan_col).value)
        if not employee_cd or not jan:
            continue
        employee_names.setdefault(employee_cd, str(ws.cell(row, name_col).value or ""))
        counts[employee_cd][jan] += 1
        rows.append(
            {
                "employee_cd": employee_cd,
                "jan": jan,
                "values": [ws.cell(row, col).value for col in range(1, len(headers) + 1)],
            }
        )
    return employee_names, counts, rows, [str(header or "") for header in headers]


def read_aladdin(ws) -> tuple[dict[str, Counter[str]], list[dict[str, Any]], dict[str, str]]:
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    records: list[dict[str, Any]] = []
    employee_names: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        employee_text = clean_aladdin_text(ws.cell(row, ALADDIN_EMPLOYEE_COL).value)
        employee_digits = normalize_digits(employee_text)
        jan = normalize_jan(ws.cell(row, ALADDIN_JAN_COL).value)
        if employee_digits.startswith(("1000", "2000")):
            continue
        if len(employee_digits) < 4 or not jan:
            continue
        employee_cd = employee_digits[-4:]
        counts[employee_cd][jan] += 1
        customer_name = clean_aladdin_text(ws.cell(row, 2).value)
        employee_names.setdefault(employee_cd, customer_name)
        cost = parse_number(ws.cell(row, ALADDIN_COST_COL).value)
        wholesale_add = parse_number(ws.cell(row, ALADDIN_WHOLESALE_ADD_COL).value)
        year, month = parse_yyyymmdd(ws.cell(row, 3).value)
        records.append(
            {
                "year": year,
                "month": month,
                "employee_cd": employee_cd,
                "customer_name": customer_name,
                "jan": jan,
                "product_code": clean_aladdin_text(ws.cell(row, ALADDIN_PRODUCT_CODE_COL).value),
                "product_name": clean_aladdin_text(ws.cell(row, ALADDIN_PRODUCT_NAME_COL).value),
                "color": clean_aladdin_text(ws.cell(row, ALADDIN_COLOR_COL).value),
                "size": clean_aladdin_text(ws.cell(row, ALADDIN_SIZE_COL).value),
                "cost": cost,
                "wholesale": cost + wholesale_add,
                "source_row": row,
            }
        )
    return counts, records, employee_names


def replace_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_pair_sheet(
    wb,
    summary_headers: list[str],
    summary_rows: list[dict[str, Any]],
    aladdin_records: list[dict[str, Any]],
    summary_counts: dict[str, Counter[str]],
) -> None:
    ws = replace_sheet(wb, PAIR_SHEET)
    headers = [
        *summary_headers,
        "アラジン/商品コード",
        "アラジン/商品名",
        "アラジン/カラー",
        "アラジン/サイズ",
        "アラジン/原価",
        "アラジン/卸価格",
        "SA照合判定",
    ]

    grouped: dict[tuple[str, str], deque[dict[str, Any]]] = defaultdict(deque)
    for record in aladdin_records:
        grouped[(record["employee_cd"], record["jan"])].append(record)

    output_rows: list[list[Any]] = []
    for summary_row in summary_rows:
        key = (summary_row["employee_cd"], summary_row["jan"])
        match = grouped[key].popleft() if grouped[key] else None
        if match:
            aladdin_values = [
                match["product_code"],
                match["product_name"],
                match["color"],
                match["size"],
                match["cost"],
                match["wholesale"],
                "一致",
            ]
        else:
            aladdin_values = ["", "", "", "", "", "", "アラジン不足"]
        output_rows.append([*summary_row["values"], *aladdin_values])

    for (employee_cd, _jan), records in sorted(grouped.items()):
        for record in records:
            summary_values = [""] * len(summary_headers)
            summary_values[0:4] = [
                record["year"],
                record["month"],
                record["employee_cd"],
                record["customer_name"],
            ]
            output_rows.append(
                [
                    *summary_values,
                    record["product_code"],
                    record["product_name"],
                    record["color"],
                    record["size"],
                    record["cost"],
                    record["wholesale"],
                    "アラジン過剰",
                ]
            )

    def sort_key(row: list[Any]) -> tuple[int, str]:
        employee_digits = normalize_digits(row[2])
        return (int(employee_digits) if employee_digits else 99999999, str(row[3] or ""))

    ws.append(headers)
    for row in sorted(output_rows, key=sort_key):
        ws.append(row)

    style_sheet(ws)
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=11):
        for cell in row:
            cell.number_format = '#,##0'
    aladdin_cost_col = len(summary_headers) + 5
    for row in ws.iter_rows(min_row=2, min_col=aladdin_cost_col, max_col=aladdin_cost_col + 1):
        for cell in row:
            cell.number_format = '#,##0'


def write_results(
    wb,
    employee_names: dict[str, str],
    summary_counts: dict[str, Counter[str]],
    aladdin_counts: dict[str, Counter[str]],
) -> int:
    result_ws = replace_sheet(wb, RESULT_SHEET)
    detail_ws = replace_sheet(wb, DETAIL_SHEET)

    result_headers = [
        "社員CD",
        "氏名",
        "集計表点数",
        "アラジン点数",
        "差異点数",
        "不足JAN数",
        "過剰JAN数",
        "判定",
    ]
    detail_headers = [
        "社員CD",
        "氏名",
        "JANコード",
        "集計表点数",
        "アラジン点数",
        "差異",
        "差異区分",
    ]
    result_ws.append(result_headers)
    detail_ws.append(detail_headers)

    mismatch_employee_count = 0
    all_employee_codes = sorted(set(summary_counts) | set(aladdin_counts))
    for employee_cd in all_employee_codes:
        summary = summary_counts.get(employee_cd, Counter())
        aladdin = aladdin_counts.get(employee_cd, Counter())
        jan_codes = sorted(set(summary) | set(aladdin))
        detail_rows = []
        shortage = 0
        excess = 0
        for jan in jan_codes:
            summary_qty = summary.get(jan, 0)
            aladdin_qty = aladdin.get(jan, 0)
            diff = aladdin_qty - summary_qty
            if diff == 0:
                continue
            if diff < 0:
                shortage += abs(diff)
                diff_type = "アラジン不足"
            else:
                excess += diff
                diff_type = "アラジン過剰"
            detail_rows.append(
                [
                    employee_cd,
                    employee_names.get(employee_cd, ""),
                    jan,
                    summary_qty,
                    aladdin_qty,
                    diff,
                    diff_type,
                ]
            )

        if detail_rows:
            mismatch_employee_count += 1
            summary_total = sum(summary.values())
            aladdin_total = sum(aladdin.values())
            result_ws.append(
                [
                    employee_cd,
                    employee_names.get(employee_cd, ""),
                    summary_total,
                    aladdin_total,
                    aladdin_total - summary_total,
                    shortage,
                    excess,
                    "不一致",
                ]
            )
            for row in detail_rows:
                detail_ws.append(row)

    style_sheet(result_ws)
    style_sheet(detail_ws)
    if mismatch_employee_count == 0:
        result_ws.append(["-", "-", 0, 0, 0, 0, 0, "全員一致"])
    return mismatch_employee_count


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="7C2D12")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row, column).value or "")) for row in range(1, min(ws.max_row, 100) + 1))
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 10), 24)


def compare(path: Path) -> tuple[Path, int]:
    wb = load_workbook(path)
    if SUMMARY_SHEET not in wb.sheetnames:
        raise ValueError(f"{SUMMARY_SHEET} シートがありません。")
    if ALADDIN_SHEET not in wb.sheetnames:
        raise ValueError(f"{ALADDIN_SHEET} シートがありません。")

    employee_names, summary_counts, summary_rows, summary_headers = read_summary(wb[SUMMARY_SHEET])
    aladdin_counts, aladdin_records, aladdin_employee_names = read_aladdin(wb[ALADDIN_SHEET])
    employee_names = {**aladdin_employee_names, **employee_names}
    write_pair_sheet(wb, summary_headers, summary_rows, aladdin_records, summary_counts)
    mismatch_count = write_results(wb, employee_names, summary_counts, aladdin_counts)
    output_path = compared_output_path(path)
    wb.save(output_path)
    return output_path, mismatch_count


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_workbook()
    output, mismatch_count = compare(path)
    print(f"output={output}")
    print(f"mismatch_employees={mismatch_count}")


if __name__ == "__main__":
    main()
