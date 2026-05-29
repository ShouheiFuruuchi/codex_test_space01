from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path
from typing import Any

import cv2
import numpy as np
import pypdfium2 as pdfium
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidocr_onnxruntime import RapidOCR

from backend import employee_from_file_name, normalize_barcode, scan_pdf


ROOT = Path(__file__).resolve().parent
PDF_DIR = ROOT / "社販表サンプル"
MASTER_PATH = ROOT / "商品マスタ" / "商品マスタ.csv"
OUTPUT_DIR = ROOT / "OUTPUT"

HEADERS = [
    "年度",
    "月",
    "社員CD",
    "氏名",
    "商品コード",
    "商品名",
    "JANコード",
    "カラー",
    "サイズ",
    "原価",
    "卸価格",
    "未照合判定",
]


def parse_number(value: Any) -> float:
    text = str(value or "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def load_product_master() -> dict[str, dict[str, Any]]:
    products: dict[str, dict[str, Any]] = {}
    with MASTER_PATH.open("r", encoding="cp932", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) < 19:
                continue
            jan = normalize_barcode(row[1])
            if not jan:
                continue
            products[jan] = {
                "product_code": row[0],
                "jan": jan,
                "name": row[2],
                "cost": parse_number(row[11]),
                "color": row[13],
                "size": row[15],
                "wholesale": parse_number(row[18]),
            }
    return products


def infer_default_year(paths: list[Path]) -> int:
    years = []
    for path in paths:
        for match in re.findall(r"(\d{4})年", path.name):
            year = int(match)
            if 2000 <= year <= 2100:
                years.append(year)
    return Counter(years).most_common(1)[0][0] if years else 2026


def infer_year_month(path: Path, default_year: int) -> tuple[int, int]:
    year_match = re.search(r"(\d{4})年", path.name)
    month_match = re.search(r"(\d{1,2})月", path.name)
    year = int(year_match.group(1)) if year_match else default_year
    if not 2000 <= year <= 2100:
        year = default_year
    month = int(month_match.group(1)) if month_match else 0
    return year, month


def render_header_image(path: Path):
    pdf = pdfium.PdfDocument(str(path))
    try:
        image = pdf[0].render(scale=4).to_pil()
    finally:
        pdf.close()

    if image.width > image.height:
        image = image.rotate(90, expand=True, fillcolor="white")

    width, height = image.size
    return image.crop((int(width * 0.42), int(height * 0.035), int(width * 0.98), int(height * 0.13)))


def clean_ocr_name(text: str) -> str:
    return re.sub(r"[^\w一-龥ぁ-んァ-ヶー々〆〤]", "", text or "")


def extract_staff_info(path: Path, ocr: RapidOCR) -> dict[str, str]:
    fallback_name = employee_from_file_name(path.name)
    crop = render_header_image(path)
    arr = cv2.cvtColor(np.array(crop), cv2.COLOR_RGB2BGR)
    result, _ = ocr(arr)
    texts = [item[1] for item in result] if result else []

    staff_cd = ""
    for text in texts:
        match = re.search(r"\b(\d{4})\b", text)
        if match:
            staff_cd = match.group(1)
            break

    name = fallback_name
    if staff_cd:
        for index, text in enumerate(texts):
            if staff_cd in text:
                candidates = [clean_ocr_name(item) for item in texts[index + 1 : index + 4]]
                candidates = [
                    item
                    for item in candidates
                    if item
                    and not re.search(r"日付|確認|貼付|社員|氏名|PC|入力|必|タグ|夕|者|\d", item)
                ]
                if candidates:
                    name = "".join(candidates[:2])
                    break

    if fallback_name and fallback_name not in name:
        name = fallback_name
    return {"staff_cd": staff_cd, "name": name, "ocr_text": " / ".join(texts)}


def build_rows() -> tuple[list[list[Any]], int, int]:
    pdf_paths = sorted(PDF_DIR.glob("*.pdf"), key=lambda p: p.name)
    default_year = infer_default_year(pdf_paths)
    months = []
    for path in pdf_paths:
        _, month = infer_year_month(path, default_year)
        if month:
            months.append(month)
    output_month = Counter(months).most_common(1)[0][0] if months else 0
    products = load_product_master()
    ocr = RapidOCR()
    rows: list[list[Any]] = []

    for path in pdf_paths:
        year, month = infer_year_month(path, default_year)
        staff = extract_staff_info(path, ocr)
        report = scan_pdf(path, path.name)

        for item in report["items"]:
            product = products.get(item["barcode"])
            quantity = int(item.get("quantity") or 1)
            for _ in range(quantity):
                rows.append(
                    [
                        year,
                        month,
                        staff["staff_cd"],
                        staff["name"],
                        product["product_code"] if product else "",
                        product["name"] if product else "",
                        item["barcode"],
                        product["color"] if product else "",
                        product["size"] if product else "",
                        product["cost"] if product else "",
                        product["wholesale"] if product else "",
                        "" if product else "未照合",
                    ]
                )
    return rows, default_year, output_month


def write_excel(rows: list[list[Any]], year: int, month: int) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"社販集計表_{year}_{month:02d}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "社販集計表"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [10, 8, 12, 16, 16, 34, 18, 14, 12, 12, 12, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=11):
        for cell in row:
            cell.number_format = '#,##0'

    summary = wb.create_sheet("社員別集計")
    summary_headers = ["年度", "月", "社員CD", "氏名", "点数", "原価合計", "卸価格合計", "未照合点数"]
    summary.append(summary_headers)
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (row[0], row[1], row[2], row[3])
        item = grouped.setdefault(key, {"qty": 0, "cost": 0.0, "wholesale": 0.0, "unmatched": 0})
        item["qty"] += 1
        item["cost"] += row[9] if isinstance(row[9], (int, float)) else 0
        item["wholesale"] += row[10] if isinstance(row[10], (int, float)) else 0
        item["unmatched"] += 1 if row[11] else 0

    for key, value in sorted(grouped.items(), key=lambda kv: (kv[0][1], kv[0][3])):
        summary.append([*key, value["qty"], value["cost"], value["wholesale"], value["unmatched"]])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate([10, 8, 12, 16, 10, 14, 14, 14], start=1):
        summary.column_dimensions[get_column_letter(index)].width = width
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    for row in summary.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '#,##0'

    wb.create_sheet("アラジンデータ")

    wb.save(output_path)
    return output_path


def main() -> None:
    rows, year, month = build_rows()
    output = write_excel(rows, year, month)
    unmatched = sum(1 for row in rows if row[-1])
    print(f"rows={len(rows)} unmatched={unmatched} output={output}")


if __name__ == "__main__":
    main()
